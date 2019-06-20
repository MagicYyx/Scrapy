#!/usr/bin/env python
# -*- coding:utf-8 -*-
# import pymongo
# import re
import cpca
import pandas as pd
from openpyxl import Workbook

# myClient = pymongo.MongoClient('localhost')
# myDB = myClient['weibo']
# myCol = myDB['中国地震台网速报']

# x = myCol.find()
# pattern = re.compile('.*正式测定.*(\d{2}月\d{2}日\d{2}时\d{2}分)在(.*?)(（.*）)发生(.*?)级.*?地震.*震源深度(.*?)千米')
# tmp_list = []
# for i in range(2775): #2775, 2849
#     lst = pattern.findall(x[i]['txt'])
#     if len(lst) != 0:
#         print(x[i]['time'], ' ', lst)
#         tmp_list.append(lst[0][1])
# df = cpca.transform(lst)

data = pd.read_excel(r'C:\Users\Administrator\Desktop\eqList.xlsx')
lst1 = list(data['发震时刻'])
lst2 = list(data['震级(M)'])
lst3 = list(data['深度(千米)'])
lst4 = list(data['参考位置'])

df = cpca.transform(lst4)
lst5 = list(df['省'])
lst6 = list(df['市'])
lst7 = list(df['区'])
lst8 = list(df['地址'])
wb = Workbook()
ws = wb.active
ws.cell(1, 1, '发震时刻')
ws.cell(1, 2, '震级(M)')
ws.cell(1, 3, '深度(千米)')
ws.cell(1, 4, '参考位置')
ws.cell(1, 5, '省')
ws.cell(1, 6, '市')
ws.cell(1, 7, '区')
ws.cell(1, 8, '地址')
for r in range(len(lst1)):
    ws.cell(r + 2, 1, lst1[r])
    ws.cell(r + 2, 2, lst2[r])
    ws.cell(r + 2, 3, lst3[r])
    ws.cell(r + 2, 4, lst4[r])
    ws.cell(r + 2, 5, lst5[r])
    ws.cell(r + 2, 6, lst6[r])
    ws.cell(r + 2, 7, lst7[r])
    ws.cell(r + 2, 8, lst8[r])
save_path = r'C:\Users\Administrator\Desktop\result.xlsx'
wb.save(save_path)

# print(df.head())
